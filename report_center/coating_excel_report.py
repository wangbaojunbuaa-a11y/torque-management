from __future__ import annotations

from datetime import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report_center.excel_report import safe_filename
from report_center.models import CoatingRecordSummary, MesPart


class CoatingExcelReportWriter:
    def write(
        self,
        staging_report_dir: str,
        record: CoatingRecordSummary,
        product_serial_no: str,
        igbt_parts: list[MesPart],
    ) -> str:
        os.makedirs(staging_report_dir, exist_ok=True)
        out_path = os.path.join(
            staging_report_dir,
            f"{safe_filename(product_serial_no.rstrip('%'))}-涂敷记录表.xlsx",
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "涂敷记录"
        ws.merge_cells("A1:H1")
        ws["A1"] = f"{product_serial_no} 涂敷记录表"
        ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        info_rows = [
            ("产品序列号", product_serial_no, "水冷基板条码", record.plate_sn),
            ("产线", f"{record.line_code} {record.line_name}", "涂敷时间", record.recorded_at),
            ("作业人员", f"{record.operator_name} ({record.operator_work_no})", "协作人员", self._assistant_text(record)),
            ("备注", record.note, "报表生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for offset, row in enumerate(info_rows, start=3):
            ws.cell(offset, 1, row[0])
            ws.cell(offset, 2, row[1])
            ws.cell(offset, 4, row[2])
            ws.cell(offset, 5, row[3])
            ws.merge_cells(start_row=offset, start_column=2, end_row=offset, end_column=3)
            ws.merge_cells(start_row=offset, start_column=5, end_row=offset, end_column=8)

        table_row = 9
        headers = ["序号", "IGBT序列号", "物料编码"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(table_row, col, header)
            cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="70AD47")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for index, part in enumerate(igbt_parts, start=1):
            ws.cell(table_row + index, 1, index)
            ws.cell(table_row + index, 2, part.barcode)
            ws.cell(table_row + index, 3, part.code)

        self._format(ws)
        wb.save(out_path)
        return out_path

    def _assistant_text(self, record: CoatingRecordSummary) -> str:
        if not record.assistant_work_no:
            return ""
        return f"{record.assistant_name} ({record.assistant_work_no})"

    def _format(self, ws) -> None:
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = cell.font.copy(name="Microsoft YaHei")
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "center",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = border
        widths = [10, 28, 18, 16, 28, 18, 18, 18]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 24
        ws.freeze_panes = "A10"
