from __future__ import annotations

from copy import copy
from datetime import datetime
import os
from pathlib import Path
import sys

from openpyxl import load_workbook

from report_center.excel_report import safe_filename
from report_center.models import CoatingRecordSummary, MesPart, WorkpieceSummary


TEMPLATE_FILENAME = "涂敷拧紧记录表-模板.xlsx"
COATING_SHEET_NAME = "功率器件涂敷记录表"
TORQUE_SHEET_NAME = "功率器件拧紧记录表"
MANUAL_DETECTION_TEXT = "手动检测，首件合格"
DEFAULT_TIGHTENING_STATION = "拧紧工作站"
TORQUE_DATA_FIRST_ROW = 7
TORQUE_DATA_LAST_TEMPLATE_ROW = 21
TORQUE_ACTUAL_FIRST_ROW = TORQUE_DATA_FIRST_ROW + 1
TORQUE_TEMPLATE_ACTUAL_CAPACITY = TORQUE_DATA_LAST_TEMPLATE_ROW - TORQUE_ACTUAL_FIRST_ROW + 1
TORQUE_TEMPLATE_LAST_ROW = 29


def split_product_serial(product_serial_no: str) -> tuple[str, str]:
    parts = product_serial_no.split("%")
    material_no = parts[1] if len(parts) > 2 else ""
    serial_no = parts[2] if len(parts) > 3 else product_serial_no.rstrip("%")
    return material_no, serial_no


class CombinedExcelReportWriter:
    def write_partial(
        self,
        output_dir: str,
        base_barcode: str,
        product_serial_no: str = "",
        coating: CoatingRecordSummary | None = None,
        workpiece: WorkpieceSummary | None = None,
        igbt_parts: list[MesPart] | None = None,
    ) -> str:
        if coating is None and workpiece is None:
            raise ValueError("未找到可用于重新生成的涂敷或拧紧记录")
        os.makedirs(output_dir, exist_ok=True)
        wb = load_workbook(self._template_path())
        if coating is not None:
            self._write_coating_sheet(
                wb[COATING_SHEET_NAME],
                coating,
                product_serial_no,
                workpiece,
                igbt_parts or [],
            )
        if workpiece is not None:
            self._write_torque_sheet(wb[TORQUE_SHEET_NAME], workpiece, product_serial_no, coating)
        label = safe_filename(product_serial_no.rstrip("%") or base_barcode)
        out_path = os.path.join(output_dir, f"{label}-涂敷拧紧记录表.xlsx")
        if os.path.exists(out_path):
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = os.path.join(output_dir, f"{label}-涂敷拧紧记录表_{stamp}.xlsx")
        wb.save(out_path)
        return out_path

    def write(
        self,
        staging_report_dir: str,
        coating: CoatingRecordSummary,
        workpiece: WorkpieceSummary,
        product_serial_no: str,
        igbt_parts: list[MesPart] | None = None,
    ) -> str:
        os.makedirs(staging_report_dir, exist_ok=True)
        out_path = os.path.join(
            staging_report_dir,
            f"{safe_filename(product_serial_no.rstrip('%'))}-涂敷拧紧记录表.xlsx",
        )

        wb = load_workbook(self._template_path())
        coating_ws = wb[COATING_SHEET_NAME]
        torque_ws = wb[TORQUE_SHEET_NAME]

        self._write_coating_sheet(coating_ws, coating, product_serial_no, workpiece, igbt_parts or [])
        self._write_torque_sheet(torque_ws, workpiece, product_serial_no, coating)

        wb.save(out_path)
        return out_path

    def _template_path(self) -> str:
        candidates = []
        env_path = os.environ.get("TORQUE_REPORT_TEMPLATE")
        if env_path:
            candidates.append(Path(env_path))
        candidates.extend(
            [
                Path.cwd() / TEMPLATE_FILENAME,
                Path(__file__).resolve().parents[1] / TEMPLATE_FILENAME,
                Path(getattr(sys, "_MEIPASS", "")) / TEMPLATE_FILENAME,
                Path(sys.executable).resolve().parent / TEMPLATE_FILENAME,
            ]
        )
        for path in candidates:
            if path and path.exists():
                return str(path)
        searched = "\n".join(str(path) for path in candidates if path)
        raise FileNotFoundError(f"未找到报表模板 {TEMPLATE_FILENAME}，已查找:\n{searched}")

    def _write_coating_sheet(
        self,
        ws,
        record: CoatingRecordSummary,
        product_serial_no: str,
        workpiece: WorkpieceSummary | None = None,
        igbt_parts: list[MesPart] | None = None,
    ) -> None:
        material_no, serial_no = split_product_serial(product_serial_no)
        coating_date, _ = self._report_dates(record, workpiece)
        parts = igbt_parts or []
        self._ensure_coating_rows(ws, len(parts))

        ws["B2"] = material_no
        ws["E2"] = serial_no
        ws["B3"] = record.plate_sn
        ws["B4"] = self._coating_operator_text(record)
        ws["E4"] = coating_date
        ws["F5"] = self._value(record, "grease_batch_no")
        ws["D6"] = self._value(record, "coating_method")
        ws["F6"] = self._value(record, "grease_open_date")
        ws["E7"] = MANUAL_DETECTION_TEXT
        self._fill_igbt_rows(ws, parts)

    def _write_torque_sheet(
        self,
        ws,
        workpiece: WorkpieceSummary,
        product_serial_no: str,
        coating: CoatingRecordSummary | None = None,
    ) -> None:
        material_no, serial_no = split_product_serial(product_serial_no)
        records = list(workpiece.records)
        self._ensure_torque_rows(ws, len(records))
        first_record = records[0] if records else None

        ws["B2"] = material_no
        ws["F2"] = serial_no
        ws["B3"] = workpiece.base_barcode
        ws["F3"] = DEFAULT_TIGHTENING_STATION
        ws["B4"] = self._rest_time(workpiece)
        ws["F4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A7"] = 1
        ws["B7"] = 1
        ws["C7"] = "0.5Nm"
        ws["D7"] = "手动拧紧"
        ws["F7"] = self._tightening_date_for_report(first_record.tightened_at, coating, workpiece) if first_record else ""
        ws["G7"] = first_record.operator_name if first_record else ""

        for index, record in enumerate(records, start=1):
            row_no = 7 + index
            values = [
                index + 1,
                record.round_no,
                self._torque_text(record.set_torque),
                record.actual_torque,
                record.actual_angle,
                self._tightening_time_for_report(record.tightened_at, coating, workpiece),
                record.operator_name,
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row_no, col, value)
            ws.cell(row_no, 4).number_format = "0.000"
            ws.cell(row_no, 5).number_format = "0.000"
        extra_rows = max(0, len(records) - TORQUE_TEMPLATE_ACTUAL_CAPACITY)
        self._clear_unused_torque_rows(
            ws,
            TORQUE_ACTUAL_FIRST_ROW + len(records),
            TORQUE_TEMPLATE_LAST_ROW + extra_rows,
        )

    def _ensure_coating_rows(self, ws, igbt_count: int) -> None:
        extra_rows = max(0, igbt_count - 10)
        if extra_rows <= 0:
            return
        insert_at = 19
        note_merge = "A19:F19"
        if note_merge in {str(item) for item in ws.merged_cells.ranges}:
            ws.unmerge_cells(note_merge)
        ws.insert_rows(insert_at, extra_rows)
        for offset in range(extra_rows):
            target_row = insert_at + offset
            self._copy_row_style(ws, 18, target_row, 6)
            ws.merge_cells(start_row=target_row, start_column=2, end_row=target_row, end_column=4)
            ws.merge_cells(start_row=target_row, start_column=5, end_row=target_row, end_column=6)
        note_row = insert_at + extra_rows
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)

    def _fill_igbt_rows(self, ws, igbt_parts: list[MesPart]) -> None:
        detail_rows = max(10, len(igbt_parts))
        for index in range(1, detail_rows + 1):
            row_no = 8 + index
            if index <= len(igbt_parts):
                ws.cell(row_no, 1, index)
                ws.cell(row_no, 2, igbt_parts[index - 1].barcode)
                ws.cell(row_no, 5, MANUAL_DETECTION_TEXT)
            else:
                for col in (1, 2, 5):
                    ws.cell(row_no, col).value = None

    def _ensure_torque_rows(self, ws, record_count: int) -> None:
        extra_rows = max(0, record_count - TORQUE_TEMPLATE_ACTUAL_CAPACITY)
        if extra_rows <= 0:
            return
        insert_at = TORQUE_DATA_LAST_TEMPLATE_ROW + 1
        ws.insert_rows(insert_at, extra_rows)
        for offset in range(extra_rows):
            self._copy_row_style(ws, TORQUE_DATA_LAST_TEMPLATE_ROW, insert_at + offset, 7)

    def _clear_unused_torque_rows(self, ws, start_row: int, end_row: int) -> None:
        if start_row > end_row:
            return
        for row_no in range(start_row, end_row + 1):
            for col in range(1, 8):
                ws.cell(row_no, col).value = None

    def _copy_row_style(self, ws, source_row: int, target_row: int, max_column: int) -> None:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
        for col in range(1, max_column + 1):
            source = ws.cell(source_row, col)
            target = ws.cell(target_row, col)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)

    def _rest_time(self, workpiece: WorkpieceSummary) -> str:
        if not workpiece.round2_completed_at or not workpiece.round3_completed_at:
            return ""
        try:
            start = datetime.fromisoformat(str(workpiece.round2_completed_at).replace(" ", "T"))
            end = datetime.fromisoformat(str(workpiece.round3_completed_at).replace(" ", "T"))
        except ValueError:
            return ""
        minutes = max(0, int((end - start).total_seconds() // 60))
        hours, remainder = divmod(minutes, 60)
        if hours:
            return f"{hours}小时{remainder}分钟"
        return f"{remainder}分钟"

    def _first_round_operator(self, workpiece: WorkpieceSummary) -> str:
        for record in workpiece.records:
            if record.round_no == 2 and record.operator_name:
                return record.operator_name
        for record in workpiece.records:
            if record.operator_name:
                return record.operator_name
        return ""

    def _coating_operator_text(self, record: CoatingRecordSummary) -> str:
        names = [record.operator_name]
        if record.assistant_name:
            names.append(record.assistant_name)
        return "、".join(name for name in names if name)

    def _torque_text(self, value: float) -> str:
        text = f"{float(value):.3f}".rstrip("0").rstrip(".")
        return f"{text}Nm"

    def _report_dates(
        self,
        coating: CoatingRecordSummary,
        workpiece: WorkpieceSummary | None,
    ) -> tuple[str, str | None]:
        coating_dt = self._parse_datetime(coating.recorded_at)
        earliest_tightening = self._earliest_tightening_time(workpiece)
        if coating_dt and earliest_tightening and coating_dt > earliest_tightening:
            aligned_date = earliest_tightening.date().isoformat()
            return aligned_date, aligned_date
        if coating_dt:
            return coating_dt.date().isoformat(), None
        return str(coating.recorded_at or ""), None

    def _tightening_time_for_report(
        self,
        tightened_at: str,
        coating: CoatingRecordSummary | None,
        workpiece: WorkpieceSummary,
    ) -> str:
        if coating is None:
            return tightened_at
        _, aligned_date = self._report_dates(coating, workpiece)
        if not aligned_date:
            return tightened_at
        tightened_dt = self._parse_datetime(tightened_at)
        if not tightened_dt:
            return aligned_date
        return f"{aligned_date} {tightened_dt.strftime('%H:%M:%S')}"

    def _tightening_date_for_report(
        self,
        tightened_at: str,
        coating: CoatingRecordSummary | None,
        workpiece: WorkpieceSummary,
    ) -> str:
        value = self._tightening_time_for_report(tightened_at, coating, workpiece)
        parsed = self._parse_datetime(value)
        return parsed.date().isoformat() if parsed else value[:10]

    def _earliest_tightening_time(self, workpiece: WorkpieceSummary | None) -> datetime | None:
        if workpiece is None:
            return None
        values = [
            value
            for value in (self._parse_datetime(record.tightened_at) for record in workpiece.records)
            if value is not None
        ]
        return min(values) if values else None

    def _parse_datetime(self, value: str | None) -> datetime | None:
        text = str(value or "").strip()
        if not text:
            return None
        normalized = text.replace("T", " ")
        for fmt, length in (
            ("%Y-%m-%d %H:%M:%S", 19),
            ("%Y-%m-%d %H:%M", 16),
            ("%Y-%m-%d", 10),
        ):
            try:
                return datetime.strptime(normalized[:length], fmt)
            except ValueError:
                pass
        try:
            return datetime.fromisoformat(text.replace(" ", "T"))
        except ValueError:
            return None

    def _value(self, record: CoatingRecordSummary, attr: str) -> str:
        return str(getattr(record, attr, "") or "")
