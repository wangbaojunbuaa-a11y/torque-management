from __future__ import annotations

from datetime import datetime
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report_center.models import MesPart, WorkpieceSummary


def safe_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", value).strip()
    return cleaned or "未命名"


class ExcelReportWriter:
    def write(
        self,
        report_dir: str,
        workpiece: WorkpieceSummary,
        product_serial_no: str,
        igbt_parts: list[MesPart],
    ) -> str:
        month_dir = datetime.now().strftime("%Y-%m")
        out_dir = os.path.join(report_dir, month_dir, safe_filename(workpiece.line_code))
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, f"{safe_filename(product_serial_no)}-拧紧记录表.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "拧紧记录"

        title = f"{product_serial_no} 拧紧记录表"
        ws.merge_cells("A1:K1")
        ws["A1"] = title
        ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        info_rows = [
            ("产品序列号", product_serial_no, "水冷基板条码", workpiece.base_barcode),
            ("产品类型", f"{workpiece.product_code} {workpiece.product_name}", "产线", f"{workpiece.line_code} {workpiece.line_name}"),
            ("第二次OK数量", workpiece.round2_ok, "第三次OK数量", workpiece.round3_ok),
            ("第二次完成时间", workpiece.round2_completed_at or "", "第三次完成时间", workpiece.round3_completed_at or ""),
            ("报表生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "要求OK数量/轮", workpiece.expected_count),
        ]
        start = 3
        for offset, row in enumerate(info_rows):
            excel_row = start + offset
            ws.cell(excel_row, 1, row[0])
            ws.cell(excel_row, 2, row[1])
            ws.cell(excel_row, 4, row[2])
            ws.cell(excel_row, 5, row[3])
            ws.merge_cells(start_row=excel_row, start_column=2, end_row=excel_row, end_column=3)
            ws.merge_cells(start_row=excel_row, start_column=5, end_row=excel_row, end_column=8)

        table_row = start + len(info_rows) + 2
        headers = [
            "序号",
            "轮次",
            "程序号",
            "目标扭矩",
            "拧紧扭矩",
            "拧紧角度",
            "拧紧时间",
            "作业人员姓名",
            "作业人员工号",
            "结果",
            "水冷基板条码",
        ]
        ws.append([])
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(table_row, col, header)
            cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for index, record in enumerate(workpiece.records, start=1):
            row_no = table_row + index
            values = [
                record.sequence_no,
                f"第{record.round_no}次",
                record.program_no,
                record.set_torque,
                record.actual_torque,
                record.actual_angle,
                record.tightened_at,
                record.operator_name,
                record.operator_work_no,
                record.result,
                workpiece.base_barcode,
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row_no, col, value)

        igbt_start = table_row + len(workpiece.records) + 3
        ws.cell(igbt_start, 1, "MES IGBT清单")
        ws.cell(igbt_start, 1).font = Font(name="Microsoft YaHei", bold=True)
        ws.cell(igbt_start + 1, 1, "IGBT序列号")
        ws.cell(igbt_start + 1, 2, "物料编码")
        for col in (1, 2):
            cell = ws.cell(igbt_start + 1, col)
            cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="70AD47")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for index, part in enumerate(igbt_parts, start=1):
            ws.cell(igbt_start + 1 + index, 1, part.barcode)
            ws.cell(igbt_start + 1 + index, 2, part.code)

        self._apply_sheet_style(ws)
        wb.save(out_path)
        return out_path

    def _apply_sheet_style(self, ws) -> None:
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = cell.font.copy(name="Microsoft YaHei")
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "center",
                    vertical="center",
                    wrap_text=True,
                    shrink_to_fit=False,
                )
                cell.border = border

        number_columns = {"D", "E", "F"}
        for col in number_columns:
            for cell in ws[col]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.000"

        for column_cells in ws.columns:
            values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
            max_len = max((len(value) for value in values), default=8)
            width = min(max(max_len + 4, 10), 42)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width

        fixed_widths = {
            "A": 10,
            "B": 10,
            "C": 10,
            "D": 12,
            "E": 12,
            "F": 12,
            "G": 22,
            "H": 18,
            "I": 16,
            "J": 10,
            "K": 30,
        }
        for col, width in fixed_widths.items():
            ws.column_dimensions[col].width = width

        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 24
        ws.freeze_panes = "A11"
        ws.auto_filter.ref = f"A10:K{max(10, ws.max_row)}"
