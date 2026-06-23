from __future__ import annotations

from datetime import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from coating.services import CoatingRecordService


class CoatingReportService:
    def __init__(self, record_service: CoatingRecordService) -> None:
        self.record_service = record_service

    def export_records(
        self,
        output_dir: str,
        start_date: str | None = None,
        end_date: str | None = None,
    ) -> str:
        rows = self.record_service.records_between(start_date, end_date)
        if not rows:
            raise ValueError("指定范围内没有涂敷记录")
        return self._write_rows(rows, output_dir, "涂敷记录")

    def export_search_results(self, rows, output_dir: str) -> str:
        if not rows:
            raise ValueError("没有可导出的涂敷记录")
        return self._write_rows(rows, output_dir, "涂敷历史查询")

    def _write_rows(self, rows, output_dir: str, title: str) -> str:
        os.makedirs(output_dir, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_file = os.path.join(output_dir, f"{title}_{stamp}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "涂敷记录"
        ws.append([title])
        ws.merge_cells("A1:K1")
        ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])
        ws.append([])
        headers = [
            "序号",
            "水冷基板条码",
            "涂敷时间",
            "作业人员姓名",
            "作业人员工号",
            "协作人员姓名",
            "协作人员工号",
            "硅脂批次号",
            "硅脂启封日期",
            "涂敷方式",
            "备注",
        ]
        ws.append(headers)
        header_row = 5
        for cell in ws[header_row]:
            cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for index, row in enumerate(rows, start=1):
            ws.append(
                [
                    index,
                    row["plate_sn"],
                    row["recorded_at"],
                    row["operator_name"],
                    row["operator_work_no"],
                    row["assistant_name"] or "",
                    row["assistant_work_no"] or "",
                    row["grease_batch_no"] or "",
                    row["grease_open_date"] or "",
                    row["coating_method"] or "",
                    row["note"] or "",
                ]
            )

        self._format(ws)
        wb.save(out_file)
        return out_file

    def _format(self, ws) -> None:
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = cell.font.copy(name="Microsoft YaHei")
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "center",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = border
        widths = [8, 30, 22, 18, 16, 18, 16, 18, 16, 18, 36]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
        for row_index in range(1, ws.max_row + 1):
            ws.row_dimensions[row_index].height = 24
        ws.freeze_panes = "A6"
        ws.auto_filter.ref = f"A5:K{ws.max_row}"
